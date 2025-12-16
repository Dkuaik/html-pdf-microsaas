import pandas as pd
import json
from io import BytesIO
import openpyxl

def analyze_results(formato_bytes, resultados_bytes):
    # Read the Formato Excel for correct answers and classifications
    df_formato = pd.read_excel(BytesIO(formato_bytes), sheet_name='ECOEMS 202526', skiprows=5)
    correct_answers = {}
    class_dict = {}
    for index, row in df_formato.iterrows():
        if pd.notna(row['ID']):
            qid = int(row['ID'])
            correct = row.iloc[8]  # Correct answer column
            correct_answers[qid] = correct.strip() if isinstance(correct, str) else str(correct).strip()
            subject = str(row['Subject']).strip() if pd.notna(row['Subject']) else ""
            class_dict[qid] = {
                "question_id": qid,
                "subject": subject,
                "topic": subject,
                "sub_topic": "",
                "sub_subtopi": "",
                "correct_answer": correct_answers[qid]
            }

    # Read student results with openpyxl
    wb = openpyxl.load_workbook(BytesIO(resultados_bytes))
    ws = wb['Sheet1']

    header_row = None
    for row in ws.iter_rows():
        if row[0].value == 'Student Name':
            header_row = row[0].row
            break

    if header_row is None:
        raise ValueError("Header row not found")

    student_data = {}
    for row in ws.iter_rows(min_row=header_row + 1):
        student_name = row[0].value
        if not student_name or student_name == 'Student Name':
            continue
        answers = []
        for col in range(4, 4 + 128):
            cell_value = row[col].value
            if cell_value:
                ans_text = str(cell_value).strip()
                answer_letter = ans_text[0] if ans_text else None
            else:
                answer_letter = None
            answers.append(answer_letter)
        student_data[student_name] = answers

    # Create student hashmap
    student_hashmap = {}
    for student, answers in student_data.items():
        student_answers = []
        for i, ans in enumerate(answers):
            qid = i + 1
            correct_letter = correct_answers.get(qid)
            is_correct = ans == correct_letter if ans and correct_letter else False
            cls = class_dict.get(qid, {})
            student_answers.append({
                "question_id": qid,
                "correct": is_correct,
                "subject": cls.get("subject", ""),
                "topic": cls.get("topic", ""),
                "sub_topic": cls.get("sub_topic", ""),
                "sub_subtopi": cls.get("sub_subtopi", "")
            })
        student_hashmap[student] = student_answers

    # Generate performance report
    performance_report = []
    for student, answers in student_hashmap.items():
        total_questions = len(answers)
        correct_count = sum(1 for a in answers if a['correct'])
        incorrect_count = total_questions - correct_count
        score_percent = (correct_count / total_questions) * 100 if total_questions > 0 else 0

        # By topic
        by_topic = {}
        for a in answers:
            topic = a['topic']
            if topic not in by_topic:
                by_topic[topic] = {'subject': a['subject'], 'correct': 0, 'incorrect': 0, 'total': 0}
            by_topic[topic]['total'] += 1
            if a['correct']:
                by_topic[topic]['correct'] += 1
            else:
                by_topic[topic]['incorrect'] += 1

        performance_report.append({
            'name': student,
            'total_correct': correct_count,
            'total_incorrect': incorrect_count,
            'score_percent': score_percent,
            'by_topic': by_topic
        })

    return student_hashmap, performance_report